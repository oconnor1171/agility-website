import openpyxl
import os

# Check which files exist
download_dir = os.path.expanduser('~\\Downloads')
revised_file = os.path.join(download_dir, 'DSP_Custom_Builders_Q1_2026_REVISED.xlsx')
original_file = os.path.join(download_dir, 'DSP_Custom_Builders_Q1_2026 (1).xlsx')

print("=== CHECKING FILES ===")
print(f"Revised file exists: {os.path.exists(revised_file)}")
print(f"Original file exists: {os.path.exists(original_file)}")
print()

if os.path.exists(revised_file):
    try:
        wb = openpyxl.load_workbook(revised_file)
        ws = wb.active
        print("=== REVISED FILE STRUCTURE ===\n")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 60:
                print(f"Row {i}: {row}")
            else:
                break
    except Exception as e:
        print(f"Error reading revised file: {e}")

print("\n" + "="*50 + "\n")

if os.path.exists(original_file):
    try:
        wb2 = openpyxl.load_workbook(original_file)
        ws2 = wb2.active
        print("=== ORIGINAL FILE STRUCTURE ===\n")
        for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
            if i <= 60:
                print(f"Row {i}: {row}")
            else:
                break
    except Exception as e:
        print(f"Error reading original file: {e}")
